from auth import s, get_token
from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink

token = get_token()


def get_user(id):
    global token
    uri = f"https://osu.ppy.sh/api/v2/users/{id}"
    header = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "Content-Type": "application/json",
    }
    return s.get(uri, headers=header).json()


def exportCommsByVotes():
    wb = Workbook()
    ws = wb.active
    repsonse = s.get("http://localhost:8080/comms/getCommsByVotes").json()
    for i in range(len(repsonse)):
        comm = repsonse[i]
        user_data = get_user(comm["user_id"])
        if 'error' in user_data:
            ws[f'A{i + 1}'] = "Deleted user"
        else:
            ws[f'A{i + 1}'] = user_data['username']
        ws[f'B{i + 1}'] = comm["msg_text"]
        ws[f'B{i + 1}'].hyperlink = Hyperlink(f"https://osu.ppy.sh/comments/{comm['id']}").ref
        ws[f'C{i + 1}'] = comm["votes"]
        print(i)


    wb.save("export/commsByVotes.xlsx")


def getBeatmapset(id):
    header = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "Content-Type": "application/json",
    }
    return s.get(f"https://osu.ppy.sh/api/v2/beatmapsets/{id}", headers=header).json()

def exportGroupedCommentableTypesByVotes():
    wb = Workbook()
    commentable_Types = ["beatmapset", "news_post", "build"]
    worksheets = ["Beatmapset", "NewsPost", "Changelog"]
    wb.create_sheet("Beatmapset")
    wb.create_sheet("NewsPost")
    wb.create_sheet("Changelog")
    del wb["Sheet"]

    for x in range(3):
        response = s.get(
            f"http://localhost:8080/comms/getGroupedCommentableTypesBy?commentable_type={commentable_Types[x]}&sort_by=votes").json()
        wb.active = wb[worksheets[x]]
        for i in range(0, len(response)):
            ctData = []
            if x == 0:
                ctData = getBeatmapsetExportData(response[i]['commentable_id'])
            elif x == 1:
                ctData = getNewsPostExportData(response[i]['commentable_id'])
            else:
                ctData = getChangelogExportData(response[i]['commentable_id'])
            ws = wb.active
            ws[f'A{i + 1}'] = ctData[0]
            ws[f'A{i + 1}'].hyperlink = ctData[1]
            ws[f'B{i + 1}'] = response[i]["votes"]
            ws[f'C{i + 1}'] = response[i]["count"]

    wb.save("export/GroupedCommentableTypesByVotes.xlsx")


def exportGroupedCommentableTypesByCount():
    wb = Workbook()
    commentable_Types = ["beatmapset", "news_post", "build"]
    worksheets = ["Beatmapset", "NewsPost", "Changelog"]
    wb.create_sheet("Beatmapset")
    wb.create_sheet("NewsPost")
    wb.create_sheet("Changelog")
    del wb["Sheet"]

    for x in range(3):
        response = s.get(f"http://localhost:8080/comms/getGroupedCommentableTypesBy?commentable_type={commentable_Types[x]}&sort_by=count").json()
        wb.active = wb[worksheets[x]]
        for i in range(0, len(response)):
            ctData = []
            if x == 0:
                ctData = getBeatmapsetExportData(response[i]['commentable_id'])
            elif x == 1:
                ctData = getNewsPostExportData(response[i]['commentable_id'])
            else:
                ctData = getChangelogExportData(response[i]['commentable_id'])
            ws = wb.active
            ws[f'A{i+1}'] = ctData[0]
            ws[f'A{i+1}'].hyperlink = ctData[1]
            ws[f'B{i+1}'] = response[i]["count"]
            ws[f'C{i+1}'] = response[i]["votes"]

    wb.save("export/GroupedCommentableTypesByCount.xlsx")

def exportPlayersByVotes():
    wb = Workbook()
    ws = wb.active
    repsonse = s.get("http://localhost:8080/players/getPlayersByVotesAlt").json()
    for i in range(len(repsonse)):
        user_data = get_user(repsonse[i]["user_id"])
        if 'error' in user_data:
            ws[f'A{i + 1}'] = "Deleted user"
            ws[f'B{i + 1}'] = repsonse[i]["comms_count"]
            ws[f'C{i + 1}'] = repsonse[i]["votes_up"]
        else:
            ws[f'A{i + 1}'] = user_data['username']
            ws[f'A{i + 1}'].hyperlink = Hyperlink(f"https://osu.ppy.sh/users/{repsonse[i]["user_id"]}").ref
            ws[f'B{i + 1}'] = repsonse[i]["comms_count"]
            ws[f'C{i + 1}'] = repsonse[i]["votes_up"]

    wb.save("export/PlayersByVotes.xlsx")

def exportPlayersByCount():
    wb = Workbook()
    ws = wb.active
    repsonse = s.get("http://localhost:8080/players/getPlayersByCountAlt").json()
    for i in range(len(repsonse)):
        user_data = get_user(repsonse[i]["user_id"])
        if 'error' in user_data:
            ws[f'A{i + 1}'] = "Deleted user"
            ws[f'B{i + 1}'] = repsonse[i]["comms_count"]
            ws[f'C{i + 1}'] = repsonse[i]["votes_up"]
        else:
            ws[f'A{i+1}'] = user_data['username']
            ws[f'A{i + 1}'].hyperlink = Hyperlink(f"https://osu.ppy.sh/users/{repsonse[i]["user_id"]}").ref
            ws[f'B{i + 1}'] = repsonse[i]["comms_count"]
            ws[f'C{i + 1}'] = repsonse[i]["votes_up"]
        print(i)
    wb.save("export/PlayersByCount.xlsx")


def getBeatmapsetExportData(beatmapset_id):
    bData = getBeatmapset(beatmapset_id)
    return [f"{bData['artist']} - {bData['title']}",
            Hyperlink(f"https://osu.ppy.sh/beatmapsets/{beatmapset_id}").ref]


def getNewsPostExportData(commentable_id):
    header = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "Content-Type": "application/json",
    }
    cData = s.get(f"https://osu.ppy.sh/api/v2/comments?commentable_type=news_post&commentable_id={commentable_id}", headers=header).json()
    newsData = cData["commentable_meta"][0]
    return [newsData["title"],
            Hyperlink(newsData["url"]).ref]


def getChangelogExportData(commentable_id):
    header = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "Content-Type": "application/json",
    }
    cData = s.get(f"https://osu.ppy.sh/api/v2/comments?commentable_type=build&commentable_id={commentable_id}", headers=header).json()
    if not 'error' in cData:
        buildData = cData["commentable_meta"][0]
        return [buildData["title"],
                Hyperlink(buildData["url"]).ref]
    return ["Not Found",
            Hyperlink("https://osu.ppy.sh").ref]


def exportCommsByDates():
    wb = Workbook()
    ws = wb.active
    for i in range(2013, 2026):
        offset = (i-2013)
        data = {
            'startDate': f"{i}-01-01",
            'endDate': f"{i+1}-01-01"
        }
        ws.merge_cells(f"A{1 + offset*6}:B{1 + offset*6}")
        ws[f"A{1 + offset*6}"] = f"{i} - {i+1}"
        response = s.get("http://localhost:8080/comms/getCommsBetweenDates", json=data).json()
        for x in range(5):
            ws[f"A{2 + x + offset*6}"] = response[x]["msg_text"]
            ws[f"B{2 + x + offset * 6}"] = response[x]["votes"]
            ws[f'A{2 + x + offset * 6}'].hyperlink = Hyperlink(f"https://osu.ppy.sh/comments/{response[x]['id']}").ref
    wb.save("export/CommsByDates.xlsx")


#exportCommsByVotes()
#exportGroupedCommentableTypesByVotes()
#exportGroupedCommentableTypesByCount()
#exportCommsByDates()
exportPlayersByCount()